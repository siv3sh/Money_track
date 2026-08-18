"""
Parse bank statement CSV / TSV / pasted text into Transaction dicts.

Supported CSV headers (case-insensitive, flexible aliases):
  date, description/narration/remarks, amount / debit+credit, type optional
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any, Optional

from parser import ALLOWED_BANKS, Transaction, _detect_bank, _detect_card_type, parse_sms

DATE_FORMATS = (
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%d-%m-%y",
    "%d/%m/%y",
    "%Y/%m/%d",
    "%d %b %Y",
    "%d-%b-%Y",
    "%d %B %Y",
    "%b %d, %Y",
)

HEADER_ALIASES = {
    "date": {"date", "txn date", "transaction date", "value date", "tran date", "posted"},
    "description": {
        "description",
        "narration",
        "remarks",
        "particulars",
        "details",
        "merchant",
        "payee",
        "transaction details",
    },
    "amount": {"amount", "txn amount", "transaction amount", "amt"},
    "debit": {"debit", "withdrawal", "dr", "debit amount", "withdrawals"},
    "credit": {"credit", "deposit", "cr", "credit amount", "deposits"},
    "type": {"type", "txn type", "transaction type", "dr/cr", "credit/debit"},
    "balance": {"balance", "closing balance", "available balance", "bal"},
}

INVESTMENT_HINTS = (
    "zerodha",
    "groww",
    "upstox",
    "kuvera",
    "indmoney",
    "mutual fund",
    "mutualfund",
    "sip ",
    " sip",
    "nse",
    "bse",
    "cdsl",
    "nsdl",
    "clearing corporation",
    "indian clearing",
    "stocksip",
    "digital gold",
    "nps ",
    "ppf",
    "broker",
    "demat",
)

CREDIT_CARD_HINTS = (
    "credit card",
    "cc xx",
    "card xx",
    "avl limit",
    "available limit",
    "spent using",
    "spent on your",
    "icici bank card",
    "purchase on card",
)


def _norm_header(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def detect_statement_profile(
    sample: str,
    filename: str = "",
) -> dict[str, Any]:
    """
    Infer bank + product (savings vs ICICI credit card) from filename / header text.
    Only Federal Bank and ICICI Bank are allowed.
    """
    blob = f"{filename}\n{sample[:12000]}".lower()
    name = (filename or "").lower()

    scores = {"Federal Bank": 0, "ICICI Bank": 0}
    if "federal" in blob or "fedbnk" in blob or "fedbk" in blob or "fed-" in name:
        scores["Federal Bank"] += 5
    if "icici" in blob or "icicib" in blob:
        scores["ICICI Bank"] += 5

    is_credit_card = any(h in blob for h in CREDIT_CARD_HINTS) or (
        "credit" in name and "card" in name
    ) or bool(re.search(r"\bcc\b|\bvisa\b|\bmastercard\b|\brupay\b", blob))

    if is_credit_card:
        # Credit-card statements in this household are ICICI
        scores["ICICI Bank"] += 4

    # Federal savings account cues
    if any(k in blob for k in ("savings a/c", "savings account", "a/c no", "account statement")):
        if scores["Federal Bank"] >= scores["ICICI Bank"]:
            scores["Federal Bank"] += 1

    bank = max(scores, key=scores.get)
    if scores[bank] <= 0:
        # Default: Federal is the primary bank account
        bank = "Federal Bank"
        confidence = "low"
    elif scores[bank] >= 5:
        confidence = "high"
    else:
        confidence = "medium"

    if bank not in ALLOWED_BANKS:
        bank = "Federal Bank"

    product = "credit_card" if (is_credit_card and bank == "ICICI Bank") else "bank_account"
    return {
        "bank": bank,
        "product": product,
        "default_card_type": "credit_card" if product == "credit_card" else "bank_account",
        "confidence": confidence,
        "is_credit_card_statement": product == "credit_card",
    }


def enrich_imported_txn(txn: dict[str, Any], *, default_card_type: str = "bank_account") -> dict[str, Any]:
    """Improve card_type / bank tagging from narration after CSV parse."""
    raw = f"{txn.get('merchant') or ''} {txn.get('raw_text') or ''}".lower()
    bank = txn.get("bank")
    if not bank or bank not in ALLOWED_BANKS:
        detected = _detect_bank(str(txn.get("sender") or ""), raw)
        if detected in ALLOWED_BANKS:
            txn["bank"] = detected

    # Card type: prefer statement product, then SMS-style heuristics on narration
    card = txn.get("card_type") or default_card_type
    if default_card_type == "credit_card":
        card = "credit_card"
    else:
        inferred = _detect_card_type(raw)
        if inferred and inferred != "bank_account":
            card = inferred
        elif any(h in raw for h in CREDIT_CARD_HINTS):
            card = "credit_card"
    txn["card_type"] = card

    # Soft investment flag for categorizer (also helps if keyword rules miss)
    if any(h in raw for h in INVESTMENT_HINTS):
        txn["_investment_hint"] = True

    # ICICI credit card last4 from narration when possible
    if card == "credit_card" and not txn.get("account_last4"):
        m = re.search(r"(?:card|xx|x{2,})\s*(\d{4})\b", raw, re.I)
        if m:
            txn["account_last4"] = m.group(1)

    return txn


def _map_headers(fieldnames: list[str]) -> dict[str, str]:
    """Map canonical name -> actual CSV header."""
    mapping: dict[str, str] = {}
    normalized = {_norm_header(f): f for f in fieldnames if f}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mapping[canonical] = normalized[alias]
                break
    return mapping


def _parse_date(raw: str) -> Optional[datetime]:
    text = (raw or "").strip()
    if not text:
        return None
    # Strip time if present
    text = re.split(r"[ T]", text, maxsplit=1)[0]
    for fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(text, fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _parse_money(raw: str) -> Optional[float]:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or text in {"-", "--", "—"}:
        return None
    # (1,234.56) accounting negative
    neg = text.startswith("(") and text.endswith(")")
    text = text.replace("₹", "").replace("INR", "").replace("Rs.", "").replace("Rs", "")
    text = text.replace(",", "").replace("(", "").replace(")", "").strip()
    text = re.sub(r"[^\d.\-]", "", text)
    if not text or text in {".", "-"}:
        return None
    try:
        value = float(text)
    except ValueError:
        return None
    if neg:
        value = -abs(value)
    return value


def _row_type(
    mapped: dict[str, str],
    row: dict[str, str],
    amount: Optional[float],
    debit: Optional[float],
    credit: Optional[float],
) -> Optional[str]:
    type_col = mapped.get("type")
    if type_col:
        t = (row.get(type_col) or "").strip().lower()
        if t in {"dr", "debit", "withdrawal", "d", "paid"}:
            return "debit"
        if t in {"cr", "credit", "deposit", "c", "received"}:
            return "credit"

    if debit is not None and debit > 0 and (credit is None or credit == 0):
        return "debit"
    if credit is not None and credit > 0 and (debit is None or debit == 0):
        return "credit"
    if amount is not None:
        if amount < 0:
            return "debit"
        if amount > 0:
            # Ambiguous single Amount column — many banks use signed or always positive.
            # Prefer description cues.
            return None
    return None


def _type_from_description(desc: str) -> Optional[str]:
    lower = desc.lower()
    if any(k in lower for k in ("debit", "withdrawal", "paid", "purchase", "upi-")):
        if "credit" not in lower[:20]:
            return "debit"
    if any(k in lower for k in ("credit", "deposit", "salary", "refund", "interest")):
        return "credit"
    return None


def parse_statement_csv(
    content: str,
    *,
    bank: Optional[str] = None,
    default_card_type: str = "bank_account",
) -> list[tuple[Transaction, datetime]]:
    """
    Returns list of (Transaction, received_at).
    """
    text = content.lstrip("\ufeff").strip()
    if not text:
        return []

    # Detect delimiter
    sample = text[:2000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        return []

    mapped = _map_headers(list(reader.fieldnames))
    if "date" not in mapped or "description" not in mapped:
        # Need at least date + description; amount via amount or debit/credit
        if "date" not in mapped:
            return []
    if "amount" not in mapped and "debit" not in mapped and "credit" not in mapped:
        return []

    bank_name = bank or None
    results: list[tuple[Transaction, datetime]] = []

    for row in reader:
        if not row:
            continue
        date_raw = row.get(mapped["date"], "") if "date" in mapped else ""
        received_at = _parse_date(date_raw)
        if received_at is None:
            continue

        desc = ""
        if "description" in mapped:
            desc = (row.get(mapped["description"]) or "").strip()
        if not desc:
            continue

        debit = _parse_money(row.get(mapped["debit"], "")) if "debit" in mapped else None
        credit = _parse_money(row.get(mapped["credit"], "")) if "credit" in mapped else None
        amount_col = _parse_money(row.get(mapped["amount"], "")) if "amount" in mapped else None

        txn_type = _row_type(mapped, row, amount_col, debit, credit)
        if txn_type is None:
            txn_type = _type_from_description(desc)
        if txn_type is None and amount_col is not None:
            # Default: positive amount column treated as debit for spend trackers
            # unless description says credit
            txn_type = "debit" if amount_col >= 0 else "credit"
            if amount_col < 0:
                amount_col = abs(amount_col)

        if txn_type == "debit":
            amount = debit if debit and debit > 0 else (abs(amount_col) if amount_col else None)
        elif txn_type == "credit":
            amount = credit if credit and credit > 0 else (abs(amount_col) if amount_col else None)
        else:
            continue

        if amount is None or amount <= 0:
            continue

        balance = None
        if "balance" in mapped:
            balance = _parse_money(row.get(mapped["balance"], ""))

        # Prefer SMS-style parse on narration when it looks like an SMS
        sms_txn = parse_sms(bank_name or "STATEMENT", desc) if bank_name else None
        if sms_txn is None and bank_name:
            # Retry with bank code-like sender for ICICI/Federal SMS embedded in statements
            sender_hint = "ICICIB" if "icici" in bank_name.lower() else "FEDBNK"
            sms_txn = parse_sms(sender_hint, desc)
        if sms_txn is not None:
            txn = sms_txn
            if bank_name:
                txn["bank"] = bank_name
            txn["raw_text"] = desc
            if balance is not None:
                txn["balance"] = balance
            if default_card_type == "credit_card":
                txn["card_type"] = "credit_card"
        else:
            merchant = desc[:60]
            # Strip leading ref codes
            merchant = re.sub(r"^(UPI|NEFT|IMPS|RTGS)[/\-\s]+", "", merchant, flags=re.I)
            txn = Transaction(
                type=txn_type,
                amount=float(amount),
                account_last4=None,
                card_type=default_card_type,
                merchant=merchant.strip() or None,
                balance=balance,
                bank=bank_name,
                raw_text=desc,
            )

        enrich_imported_txn(txn, default_card_type=default_card_type)
        results.append((txn, received_at))

    return results


def parse_statement_text(
    content: str,
    *,
    bank: Optional[str] = None,
    sender: str = "STATEMENT",
) -> list[tuple[Transaction, datetime]]:
    """
    Each non-empty line: optional leading date, then SMS-like or free text.
    Examples:
      2026-07-15 Rs.200 debited ... at SWIGGY
      15/07/2026,SWIGGY,200,DR
    """
    results: list[tuple[Transaction, datetime]] = []
    # Try CSV first if it looks like one
    if "," in content and ("date" in content.lower()[:200] or "narration" in content.lower()[:200]):
        csv_rows = parse_statement_csv(content, bank=bank)
        if csv_rows:
            return csv_rows

    date_line_re = re.compile(
        r"^(\d{1,4}[-/]\d{1,2}[-/]\d{1,4}|\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})\s*[,:\-]?\s*(.+)$"
    )

    for line in content.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        received_at = datetime.now(timezone.utc)
        body = line
        m = date_line_re.match(line)
        if m:
            parsed = _parse_date(m.group(1))
            if parsed:
                received_at = parsed
                body = m.group(2).strip()

        txn = parse_sms(sender if not bank else bank, body)
        if txn is None:
            # Minimal fallback: amount + DR/CR
            amount_m = re.search(
                r"(?:Rs\.?|INR|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
                body,
                re.I,
            )
            if not amount_m:
                continue
            amount = float(amount_m.group(1).replace(",", ""))
            lower = body.lower()
            if any(k in lower for k in (" cr", "credit", "deposit")):
                txn_type = "credit"
            elif any(k in lower for k in (" dr", "debit", "withdraw", "paid", "spent")):
                txn_type = "debit"
            else:
                continue
            txn = Transaction(
                type=txn_type,
                amount=amount,
                account_last4=None,
                card_type="bank_account",
                merchant=body[:60],
                balance=None,
                bank=bank,
                raw_text=body,
            )
        elif bank:
            txn["bank"] = bank

        results.append((txn, received_at))

    return results


def _cell_to_str(c: Any) -> str:
    if c is None:
        return ""
    if isinstance(c, datetime):
        return c.strftime("%Y-%m-%d")
    if isinstance(c, float) and c == int(c) and 20000 < c < 60000:
        # Likely Excel serial date left as float (rare with data_only)
        try:
            from datetime import timedelta

            base = datetime(1899, 12, 30, tzinfo=timezone.utc)
            return (base + timedelta(days=int(c))).strftime("%Y-%m-%d")
        except Exception:  # noqa: BLE001
            pass
    return str(c).strip()


def _rows_to_csv_text(rows: list[list[Any]]) -> str:
    """Serialize a 2D table to CSV text for reuse of parse_statement_csv."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow([_cell_to_str(c) for c in row])
    return buf.getvalue()


def _find_header_row(rows: list[list[Any]], scan: int = 40) -> int:
    """Pick the row that looks most like column headers (date + description)."""
    best_idx, best_score = 0, -1
    for i, row in enumerate(rows[:scan]):
        cells = [_norm_header(str(c)) for c in row if c is not None and str(c).strip()]
        if not cells:
            continue
        score = 0
        joined = " ".join(cells)
        for aliases in HEADER_ALIASES.values():
            if any(a in cells or a in joined for a in aliases):
                score += 1
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def parse_statement_excel(
    raw: bytes,
    *,
    bank: Optional[str] = None,
    filename: str = "",
) -> list[tuple[Transaction, datetime]]:
    """Parse .xlsx/.xls — tries every sheet and keeps the richest parse."""
    name = (filename or "").lower()
    best: list[tuple[Transaction, datetime]] = []

    def _parse_rows(rows: list[list[Any]]) -> list[tuple[Transaction, datetime]]:
        if not rows:
            return []
        header_idx = _find_header_row(rows)
        table = [r for r in rows[header_idx:] if any(c is not None and str(c).strip() for c in r)]
        if len(table) < 2:
            return []
        return parse_statement_csv(_rows_to_csv_text(table), bank=bank)

    if name.endswith(".xls") and not name.endswith(".xlsx"):
        import xlrd  # type: ignore

        book = xlrd.open_workbook(file_contents=raw)
        for i in range(book.nsheets):
            sheet = book.sheet_by_index(i)
            rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
            pairs = _parse_rows(rows)
            if len(pairs) > len(best):
                best = pairs
    else:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = [list(row) for row in ws.iter_rows(values_only=True) if row]
                pairs = _parse_rows(rows)
                if len(pairs) > len(best):
                    best = pairs
        finally:
            wb.close()

    return best


def parse_statement_pdf(
    raw: bytes,
    *,
    bank: Optional[str] = None,
) -> list[tuple[Transaction, datetime]]:
    """
    Parse PDF bank statements:
    1) extract tables → CSV path
    2) plain text line parsing
    Prefer whichever yields more unique rows (tables can be partial).
    """
    import pdfplumber

    table_pairs: list[tuple[Transaction, datetime]] = []
    text_chunks: list[str] = []

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                if not table or len(table) < 2:
                    continue
                cleaned = [
                    [("" if c is None else str(c).replace("\n", " ").strip()) for c in row]
                    for row in table
                    if row and any(c is not None and str(c).strip() for c in row)
                ]
                if len(cleaned) < 2:
                    continue
                header_idx = _find_header_row(cleaned)
                slice_rows = cleaned[header_idx:]
                pairs = parse_statement_csv(_rows_to_csv_text(slice_rows), bank=bank)
                table_pairs.extend(pairs)

            page_text = page.extract_text() or ""
            if page_text.strip():
                text_chunks.append(page_text)

    combined = "\n".join(text_chunks)
    text_pairs = parse_statement_text(combined, bank=bank) if combined.strip() else []

    if len(table_pairs) >= max(3, len(text_pairs)):
        return table_pairs
    if text_pairs:
        return text_pairs
    return table_pairs


def parse_statement(
    content: str,
    *,
    bank: Optional[str] = None,
    fmt: str = "auto",
    filename: str = "",
) -> list[dict[str, Any]]:
    """
    Unified entry for text/CSV: returns docs ready for Mongo insert.
    Auto-detects Federal vs ICICI (+ credit-card product) when bank is omitted.
    """
    profile = detect_statement_profile(content, filename)
    bank_name = bank if bank in ALLOWED_BANKS else profile["bank"]
    default_card = profile["default_card_type"]

    fmt = (fmt or "auto").lower()
    if fmt == "csv":
        pairs = parse_statement_csv(content, bank=bank_name, default_card_type=default_card)
    elif fmt == "text":
        pairs = parse_statement_text(content, bank=bank_name)
    else:
        pairs = parse_statement_csv(content, bank=bank_name, default_card_type=default_card)
        if not pairs:
            pairs = parse_statement_text(content, bank=bank_name)

    return _pairs_to_docs(pairs, default_card_type=default_card, bank=bank_name)


def parse_statement_file(
    raw: bytes,
    filename: str,
    *,
    bank: Optional[str] = None,
    fmt: str = "auto",
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """
    Parse uploaded file bytes (.csv/.txt/.tsv/.xlsx/.xls/.pdf).
    Returns (docs, detected_profile).
    """
    name = (filename or "").lower()
    fmt = (fmt or "auto").lower()

    # Peek text for bank detection
    peek = ""
    try:
        peek = raw[:8000].decode("utf-8-sig", errors="ignore")
    except Exception:  # noqa: BLE001
        peek = ""
    profile = detect_statement_profile(f"{filename}\n{peek}", filename)
    bank_name = bank if bank in ALLOWED_BANKS else profile["bank"]
    default_card = profile["default_card_type"]
    # If user forced bank but profile says credit card and bank is ICICI, keep CC
    if bank_name == "ICICI Bank" and profile["is_credit_card_statement"]:
        default_card = "credit_card"

    if name.endswith(".pdf") or fmt == "pdf":
        pairs = parse_statement_pdf(raw, bank=bank_name)
        docs = _pairs_to_docs(pairs, default_card_type=default_card, bank=bank_name)
        return docs, {**profile, "bank": bank_name, "default_card_type": default_card}

    if name.endswith(".xlsx") or name.endswith(".xls") or fmt in {"xlsx", "xls", "excel"}:
        pairs = parse_statement_excel(raw, bank=bank_name, filename=name)
        docs = _pairs_to_docs(pairs, default_card_type=default_card, bank=bank_name)
        return docs, {**profile, "bank": bank_name, "default_card_type": default_card}

    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = raw.decode("latin-1")

    profile = detect_statement_profile(content, filename)
    bank_name = bank if bank in ALLOWED_BANKS else profile["bank"]
    default_card = profile["default_card_type"]

    if name.endswith(".csv") or fmt == "csv":
        pairs = parse_statement_csv(content, bank=bank_name, default_card_type=default_card)
    elif name.endswith(".txt") or fmt == "text":
        pairs = parse_statement_text(content, bank=bank_name)
    else:
        pairs = parse_statement_csv(content, bank=bank_name, default_card_type=default_card)
        if not pairs:
            pairs = parse_statement_text(content, bank=bank_name)

    docs = _pairs_to_docs(pairs, default_card_type=default_card, bank=bank_name)
    return docs, {**profile, "bank": bank_name, "default_card_type": default_card}


def _pairs_to_docs(
    pairs: list[tuple[Transaction, datetime]],
    *,
    default_card_type: str = "bank_account",
    bank: Optional[str] = None,
) -> list[dict[str, Any]]:
    docs: list[dict[str, Any]] = []
    for txn, received_at in pairs:
        doc = {
            **txn,
            "sender": "STATEMENT",
            "received_at": received_at,
            "source": "statement",
        }
        if bank:
            doc["bank"] = bank
        enrich_imported_txn(doc, default_card_type=default_card_type)
        docs.append(doc)
    return docs