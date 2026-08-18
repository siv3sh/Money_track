"""LangGraph multi-agent import pipeline.

Nodes: extract → schema_map → categorize → validate → review
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime, timezone
from typing import Any, Literal, Optional, TypedDict

from langgraph.graph import END, StateGraph

from categorizer import CATEGORIES, apply_category, categorize
from indmoney_import import HEADER_ALIASES as IND_ALIASES
from indmoney_import import TARGET_FIELDS, suggest_mapping
from llm import LLMError, extract_json, get_provider
from merchant_label import clean_merchant_label
from rag.embeddings import embed_text
from rag.vector_store import VectorStore, get_vector_store
from statement_parser import (
    HEADER_ALIASES as STMT_ALIASES,
    detect_statement_profile,
    enrich_imported_txn,
    parse_statement_csv,
    parse_statement_excel,
    parse_statement_pdf,
    parse_statement_text,
    _find_header_row,
    _pairs_to_docs,
    _rows_to_csv_text,
)

Flow = Literal["statement", "indmoney"]
STAGE_ORDER = [
    "extracting",
    "mapping_columns",
    "categorizing",
    "validating",
    "ready_for_review",
    "complete",
]

CAT_CONFIDENCE_AUTO = float(__import__("os").getenv("IMPORT_CAT_CONFIDENCE_AUTO", "0.72"))
DOC_AUTO_THRESHOLD = float(__import__("os").getenv("RAG_DOC_MATCH_THRESHOLD", "0.62"))


class ImportGraphState(TypedDict, total=False):
    raw: bytes
    filename: str
    flow: Flow
    bank: Optional[str]
    stage: str
    steps: list[dict[str, Any]]
    warnings: list[str]
    file_kind: str
    sheets: list[dict[str, Any]]
    header_signature: str
    headers: list[str]
    raw_rows: list[dict[str, Any]]
    column_mapping: dict[str, Any]
    mapping_confidence: float
    needs_mapping_ui: bool
    profile: dict[str, Any]
    rows: list[dict[str, Any]]  # normalized candidates
    commit_rows: list[dict[str, Any]]
    review_queue: list[dict[str, Any]]
    anomalies: list[dict[str, Any]]
    review_questions: list[dict[str, Any]]
    vector_backend: str
    error: Optional[str]
    _merchant_memory: dict[str, str]
    _exact_fps: list[str]
    _soft_fps: list[str]


def _step(tool: str, status: str, detail: str, **evidence: Any) -> dict[str, Any]:
    return {"tool": tool, "status": status, "detail": detail, "evidence": evidence}


def _append_step(state: ImportGraphState, step: dict[str, Any]) -> list[dict[str, Any]]:
    steps = list(state.get("steps") or [])
    steps.append(step)
    return steps


def _header_signature(headers: list[str], sample_rows: list[list[Any]] | None = None) -> str:
    parts = ["|".join(h.strip().lower() for h in headers if h)]
    if sample_rows:
        for row in sample_rows[:2]:
            parts.append("|".join(str(c or "")[:40] for c in row[:8]))
    return "\n".join(parts)[:1500]


def _detect_kind(filename: str) -> str:
    name = (filename or "").lower()
    if name.endswith(".pdf"):
        return "pdf"
    if name.endswith((".xlsx", ".xls")):
        return "excel"
    if name.endswith((".csv", ".tsv")):
        return "csv"
    if name.endswith(".txt"):
        return "text"
    return "unknown"


def _ocr_pdf_pages(raw: bytes) -> str:
    """Best-effort OCR for scanned PDFs. Optional deps: pdf2image + pytesseract."""
    try:
        from pdf2image import convert_from_bytes  # type: ignore
        import pytesseract  # type: ignore
    except Exception:
        return ""
    try:
        images = convert_from_bytes(raw, dpi=200, first_page=1, last_page=5)
        chunks: list[str] = []
        for img in images:
            chunks.append(pytesseract.image_to_string(img) or "")
        return "\n".join(chunks)
    except Exception:  # noqa: BLE001
        return ""


def _sheet_purpose_from_headers(headers: list[str], store: VectorStore) -> str:
    sig = _header_signature(headers)
    # Known purpose patterns via embedding similarity to DocumentFormat sheet_purpose seeds
    matches = store.match_document_formats(sig, flow="indmoney", k=3, threshold=0.4)
    for m in matches:
        purpose = m.get("sheet_purpose")
        if purpose:
            return str(purpose)
    blob = " ".join(headers).lower()
    if any(k in blob for k in ("invested", "current value", "market value", "units", "scheme")):
        return "holdings"
    if any(k in blob for k in ("debit", "credit", "narration", "withdrawal")):
        return "transactions"
    return "unknown"


# ── Nodes ───────────────────────────────────────────────────────────


def extraction_agent(state: ImportGraphState) -> ImportGraphState:
    raw = state["raw"]
    filename = state.get("filename") or "upload"
    flow = state.get("flow") or "statement"
    kind = _detect_kind(filename)
    warnings = list(state.get("warnings") or [])
    store = get_vector_store()
    sheets_meta: list[dict[str, Any]] = []
    headers: list[str] = []
    header_signature = ""
    raw_rows: list[dict[str, Any]] = []
    profile: dict[str, Any] = {}
    pairs: list = []

    if kind == "pdf":
        pairs = parse_statement_pdf(raw, bank=state.get("bank"))
        # If empty, try OCR fallback for scanned PDFs
        if not pairs:
            ocr_text = _ocr_pdf_pages(raw)
            if ocr_text.strip():
                pairs = parse_statement_text(ocr_text, bank=state.get("bank"))
                warnings.append("Used OCR fallback for scanned PDF")
            else:
                warnings.append(
                    "PDF had little extractable text — install pdf2image+pytesseract for OCR"
                )
        peek = ""
        try:
            import pdfplumber

            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                peek = "\n".join((p.extract_text() or "") for p in pdf.pages[:3])[:8000]
        except Exception:  # noqa: BLE001
            peek = ""
        profile = detect_statement_profile(f"{filename}\n{peek}", filename)
        docs = _pairs_to_docs(
            pairs,
            default_card_type=profile.get("default_card_type") or "bank_account",
            bank=state.get("bank") or profile.get("bank"),
        )
        raw_rows = docs
        # Synthetic headers for DocumentFormat signature
        headers = ["date", "description", "amount", "type", "balance"]
        header_signature = _header_signature(
            headers, [[d.get("received_at"), d.get("raw_text"), d.get("amount")] for d in docs[:3]]
        )
        status = "ok" if docs else "warn"
        detail = f"PDF extract · {len(docs)} rows" + (" · OCR" if "OCR" in " ".join(warnings) else "")
    elif kind == "excel":
        # Multi-sheet: score by purpose embedding + parse richness
        best_docs: list[dict[str, Any]] = []
        best_headers: list[str] = []
        best_sig = ""
        name = filename.lower()
        try:
            if name.endswith(".xls") and not name.endswith(".xlsx"):
                import xlrd  # type: ignore

                book = xlrd.open_workbook(file_contents=raw)
                sheet_iter = [
                    (book.sheet_by_index(i).name, [
                        [book.sheet_by_index(i).cell_value(r, c) for c in range(book.sheet_by_index(i).ncols)]
                        for r in range(book.sheet_by_index(i).nrows)
                    ])
                    for i in range(book.nsheets)
                ]
            else:
                from openpyxl import load_workbook

                wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                sheet_iter = []
                try:
                    for sn in wb.sheetnames:
                        ws = wb[sn]
                        rows = [list(row) for row in ws.iter_rows(values_only=True) if row]
                        sheet_iter.append((sn, rows))
                finally:
                    wb.close()

            for sheet_name, rows in sheet_iter:
                if not rows:
                    continue
                hdr_idx = _find_header_row(rows)
                hdr = [str(c or "").strip() for c in rows[hdr_idx]]
                purpose = _sheet_purpose_from_headers(hdr, store) if flow == "indmoney" else "transactions"
                table = [r for r in rows[hdr_idx:] if any(c is not None and str(c).strip() for c in r)]
                parsed_n = 0
                docs_local: list[dict[str, Any]] = []
                if flow == "statement" and len(table) >= 2:
                    pairs_local = parse_statement_csv(_rows_to_csv_text(table), bank=state.get("bank"))
                    profile = detect_statement_profile(
                        f"{filename}\n{' '.join(hdr)}", filename
                    )
                    docs_local = _pairs_to_docs(
                        pairs_local,
                        default_card_type=profile.get("default_card_type") or "bank_account",
                        bank=state.get("bank") or profile.get("bank"),
                    )
                    parsed_n = len(docs_local)
                elif flow == "indmoney" and purpose in {"holdings", "unknown"} and len(table) >= 2:
                    # Keep as dict rows for mapping agent
                    keys = hdr
                    for r in table[1:]:
                        docs_local.append(
                            {keys[i]: r[i] if i < len(r) else None for i in range(len(keys))}
                        )
                    parsed_n = len(docs_local)

                sheets_meta.append(
                    {
                        "name": sheet_name,
                        "purpose": purpose,
                        "headers": hdr,
                        "parsed": parsed_n,
                    }
                )
                # Prefer holdings sheet for indmoney, richest for statement
                better = False
                if flow == "indmoney":
                    better = purpose == "holdings" and parsed_n >= len(best_docs)
                    if not best_docs and parsed_n:
                        better = True
                else:
                    better = parsed_n > len(best_docs)
                if better:
                    best_docs = docs_local
                    best_headers = hdr
                    best_sig = _header_signature(hdr, table[1:3] if len(table) > 1 else None)
                    raw_rows = docs_local
                    headers = hdr
                    header_signature = best_sig
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"Excel multi-sheet scan failed: {exc}")
            if flow == "statement":
                pairs = parse_statement_excel(raw, bank=state.get("bank"), filename=filename)
                profile = detect_statement_profile(filename, filename)
                raw_rows = _pairs_to_docs(
                    pairs,
                    default_card_type=profile.get("default_card_type") or "bank_account",
                    bank=state.get("bank") or profile.get("bank"),
                )
        status = "ok" if raw_rows else "warn"
        detail = f"Excel · {len(sheets_meta)} sheets · {len(raw_rows)} rows"
    else:
        # CSV / text
        try:
            content = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            content = raw.decode("latin-1")
        if flow == "statement":
            profile = detect_statement_profile(content, filename)
            bank = state.get("bank") or profile.get("bank")
            pairs = parse_statement_csv(
                content, bank=bank, default_card_type=profile.get("default_card_type") or "bank_account"
            )
            if not pairs:
                pairs = parse_statement_text(content, bank=bank)
            raw_rows = _pairs_to_docs(
                pairs,
                default_card_type=profile.get("default_card_type") or "bank_account",
                bank=bank,
            )
            # Detect headers from first line
            first = next((ln for ln in content.splitlines() if ln.strip()), "")
            headers = [h.strip() for h in re.split(r"[,|\t]", first)]
            header_signature = _header_signature(headers)
        else:
            reader = csv.DictReader(io.StringIO(content))
            headers = list(reader.fieldnames or [])
            raw_rows = [dict(r) for r in reader]
            header_signature = _header_signature(headers)
        status = "ok" if raw_rows else "warn"
        detail = f"{kind.upper()} extract · {len(raw_rows)} rows"

    if state.get("bank") and profile:
        profile = {**profile, "bank": state["bank"]}

    return {
        **state,
        "stage": "extracting",
        "file_kind": kind,
        "sheets": sheets_meta,
        "headers": headers,
        "header_signature": header_signature,
        "raw_rows": raw_rows,
        "profile": profile,
        "warnings": warnings,
        "vector_backend": store.backend,
        "steps": _append_step(
            state,
            _step(
                "extraction_agent",
                status,
                detail,
                kind=kind,
                sheets=len(sheets_meta),
                rows=len(raw_rows),
            ),
        ),
    }


def schema_mapping_agent(state: ImportGraphState) -> ImportGraphState:
    flow = state.get("flow") or "statement"
    store = get_vector_store()
    sig = state.get("header_signature") or ""
    headers = list(state.get("headers") or [])
    mapping: dict[str, Any] = {}
    confidence = 0.0
    needs_ui = False

    matches = store.match_document_formats(sig, flow=flow, k=3) if sig else []
    if matches and float(matches[0].get("similarity") or 0) >= DOC_AUTO_THRESHOLD:
        mapping = dict(matches[0].get("column_mapping") or {})
        confidence = float(matches[0].get("similarity") or 0)
        status = "ok"
        detail = f"RAG DocumentFormat match · similarity {confidence:.2f}"
    else:
        # Heuristic + optional LLM
        if flow == "indmoney":
            mapping = suggest_mapping(headers)
            # score: fraction of required fields mapped
            required = [f["key"] for f in TARGET_FIELDS if f.get("required")]
            hit = sum(1 for k in required if mapping.get(k))
            confidence = hit / max(len(required), 1) * 0.55
            if confidence < DOC_AUTO_THRESHOLD:
                try:
                    mapping = _llm_infer_mapping(headers, flow="indmoney") or mapping
                    confidence = max(confidence, 0.5)
                except Exception:  # noqa: BLE001
                    pass
            needs_ui = confidence < DOC_AUTO_THRESHOLD or not mapping.get("name")
        else:
            # Statement: map aliases into canonical roles
            mapping = _suggest_statement_mapping(headers)
            confidence = 0.7 if mapping.get("description") and mapping.get("date") else 0.4
            if confidence < DOC_AUTO_THRESHOLD and headers:
                try:
                    mapping = _llm_infer_mapping(headers, flow="statement") or mapping
                    confidence = max(confidence, 0.55)
                except Exception:  # noqa: BLE001
                    pass
            needs_ui = confidence < DOC_AUTO_THRESHOLD
        status = "ok" if not needs_ui else "warn"
        detail = (
            f"Inferred mapping · confidence {confidence:.2f}"
            + (" · needs confirmation UI" if needs_ui else "")
        )
        if matches:
            detail += f" · nearest RAG {float(matches[0].get('similarity') or 0):.2f}"

    return {
        **state,
        "stage": "mapping_columns",
        "column_mapping": mapping,
        "mapping_confidence": confidence,
        "needs_mapping_ui": needs_ui,
        "steps": _append_step(
            state,
            _step(
                "schema_mapping_agent",
                status,
                detail,
                mapping=mapping,
                confidence=confidence,
                needs_mapping_ui=needs_ui,
                rag_hits=len(matches),
            ),
        ),
    }


def _suggest_statement_mapping(headers: list[str]) -> dict[str, str | None]:
    norm = {h: re.sub(r"\s+", " ", h.strip().lower()) for h in headers}
    out: dict[str, str | None] = {k: None for k in STMT_ALIASES}
    for field, aliases in STMT_ALIASES.items():
        for original, n in norm.items():
            if n in aliases or any(a in n for a in aliases):
                out[field] = original
                break
    return out


def _llm_infer_mapping(headers: list[str], *, flow: str) -> dict[str, Any] | None:
    provider = get_provider()
    if flow == "indmoney":
        targets = [f["key"] for f in TARGET_FIELDS]
        system = (
            "Map spreadsheet headers to target fields. Return JSON object only: "
            '{ "field": "Header Name or null", ... }'
        )
        user = f"Targets: {targets}\nHeaders: {headers}"
    else:
        targets = list(STMT_ALIASES.keys())
        system = (
            "Map bank statement headers to roles date,description,amount,debit,credit,type,balance. "
            "Return JSON object only."
        )
        user = f"Targets: {targets}\nHeaders: {headers}"
    text = provider.complete(system, user, temperature=0.0)
    data = extract_json(text)
    return data if isinstance(data, dict) else None


def categorization_agent(state: ImportGraphState) -> ImportGraphState:
    flow = state.get("flow") or "statement"
    store = get_vector_store()
    memory = state.get("_merchant_memory") or {}  # injected by runner
    rows_in = list(state.get("raw_rows") or [])
    out_rows: list[dict[str, Any]] = []

    if flow == "indmoney":
        # Holdings don't need spend categories — attach asset classification lightly
        for row in rows_in:
            item = dict(row)
            item["category_confidence"] = 0.9
            item["category_source"] = "indmoney_holdings"
            out_rows.append(item)
        detail = f"INDmoney holdings · {len(out_rows)} rows (no spend categorize)"
        return {
            **state,
            "stage": "categorizing",
            "rows": out_rows,
            "steps": _append_step(
                state,
                _step("categorization_agent", "ok", detail, rows=len(out_rows)),
            ),
        }

    rag_hits = 0
    llm_hits = 0
    exact_hits = 0
    for doc in rows_in:
        item = dict(doc)
        apply_category(item, memory)
        src = str(item.get("category_source") or "")
        cat = str(item.get("category") or "Other")

        # Exact merchant memory already applied inside categorize
        if src in {"memory", "learned"} or (
            src.startswith("memory") or cat != "Other" and src not in {"other", "type_default"}
        ):
            if src in {"memory", "learned", "mcc", "keyword", "family_transfer"}:
                exact_hits += 1
                item["category_confidence"] = 0.95 if src in {"memory", "learned"} else 0.85
                out_rows.append(item)
                continue

        if cat != "Other" and src not in {"other", "type_default"}:
            item["category_confidence"] = 0.8
            out_rows.append(item)
            continue

        # RAG few-shot
        narration = f"{item.get('merchant') or ''} {item.get('raw_text') or ''}".strip()
        neighbors = store.match_merchant_knowledge(narration, k=5)
        if neighbors and float(neighbors[0].get("similarity") or 0) >= 0.55:
            item["category"] = neighbors[0]["category"]
            item["category_source"] = "rag:merchant_knowledge"
            item["category_confidence"] = float(neighbors[0].get("similarity") or 0.55)
            item["classification"] = neighbors[0].get("classification")
            rag_hits += 1
            out_rows.append(item)
            continue

        # LLM with retrieved few-shots (skip quickly if no provider)
        try:
            from llm import LLMError, provider_status

            if not provider_status().get("active"):
                item["category_confidence"] = 0.2
                out_rows.append(item)
                continue
            cat_llm, conf = _llm_categorize(narration, neighbors)
            if cat_llm:
                item["category"] = cat_llm
                item["category_source"] = "llm:rag_fewshot"
                item["category_confidence"] = conf
                llm_hits += 1
            else:
                item["category_confidence"] = 0.2
        except Exception:  # noqa: BLE001
            item["category_confidence"] = 0.15
        out_rows.append(item)

    return {
        **state,
        "stage": "categorizing",
        "rows": out_rows,
        "steps": _append_step(
            state,
            _step(
                "categorization_agent",
                "ok",
                f"exact {exact_hits} · RAG {rag_hits} · LLM {llm_hits} · total {len(out_rows)}",
                exact=exact_hits,
                rag=rag_hits,
                llm=llm_hits,
            ),
        ),
    }


def _llm_categorize(
    narration: str, neighbors: list[dict[str, Any]]
) -> tuple[str | None, float]:
    examples = "\n".join(
        f"- {n.get('description')} => {n.get('category')}" for n in neighbors[:5]
    )
    system = (
        "Classify the bank transaction into exactly one category from this list: "
        + ", ".join(CATEGORIES)
        + '. Return JSON: {"category":"...","confidence":0-1}'
    )
    user = f"Examples:\n{examples or '(none)'}\n\nTransaction:\n{narration}"
    provider = get_provider()
    text = provider.complete(system, user, temperature=0.0)
    data = extract_json(text)
    if not isinstance(data, dict):
        return None, 0.2
    cat = str(data.get("category") or "").strip()
    if cat not in CATEGORIES:
        return None, 0.2
    try:
        conf = float(data.get("confidence") or 0.55)
    except (TypeError, ValueError):
        conf = 0.55
    return cat, max(0.0, min(conf, 1.0))


def validation_dedup_agent(state: ImportGraphState) -> ImportGraphState:
    """Cross-source dedup fingerprints + anomaly flags. Existing keys injected by runner."""
    flow = state.get("flow") or "statement"
    rows = list(state.get("rows") or [])
    exact: set[str] = set(state.get("_exact_fps") or [])
    soft: set[str] = set(state.get("_soft_fps") or [])
    commit: list[dict[str, Any]] = []
    review: list[dict[str, Any]] = []
    anomalies: list[dict[str, Any]] = []
    skipped_exact = 0
    skipped_soft = 0

    amounts = [float(r.get("amount") or 0) for r in rows if r.get("amount") is not None]
    median = sorted(amounts)[len(amounts) // 2] if amounts else 0.0

    for row in rows:
        item = dict(row)
        if flow == "statement":
            fp = _fingerprint(item)
            sk = _soft_fingerprint(item)
            if fp in exact:
                skipped_exact += 1
                item["dup_status"] = "exact"
                review.append({**item, "review_reason": "exact_duplicate"})
                continue
            if sk in soft:
                skipped_soft += 1
                item["dup_status"] = "likely"
                review.append({**item, "review_reason": "likely_cross_source_duplicate"})
                continue
            exact.add(fp)
            soft.add(sk)

            # Anomalies
            amt = float(item.get("amount") or 0)
            flags: list[str] = []
            if median and amt > max(median * 8, 50000):
                flags.append("unusually_large_amount")
            received = item.get("received_at")
            if isinstance(received, datetime) and received.year < 2018:
                flags.append("malformed_or_old_date")
            conf = float(item.get("category_confidence") or 0)
            if flags:
                anomalies.append({"flags": flags, "merchant": item.get("merchant"), "amount": amt})
                item["anomaly_flags"] = flags
                item["review_reason"] = "anomaly:" + ",".join(flags)
                review.append(item)
                continue
            if conf < CAT_CONFIDENCE_AUTO or (item.get("category") or "Other") == "Other":
                item["review_reason"] = "low_category_confidence"
                review.append(item)
                continue
            commit.append(item)
        else:
            # INDmoney holdings: basic required field checks
            mapping = state.get("column_mapping") or {}
            # rows may still be raw dicts — validation happens later with apply_mapping
            commit.append(item)

    return {
        **state,
        "stage": "validating",
        "commit_rows": commit,
        "review_queue": review,
        "anomalies": anomalies,
        "steps": _append_step(
            state,
            _step(
                "validation_dedup_agent",
                "ok",
                f"commit {len(commit)} · review {len(review)} · exact_dups {skipped_exact} · soft_dups {skipped_soft}",
                commit=len(commit),
                review=len(review),
                exact_dups=skipped_exact,
                soft_dups=skipped_soft,
                anomalies=len(anomalies),
            ),
        ),
    }


def _fingerprint(doc: dict[str, Any]) -> str:
    ts = doc.get("received_at")
    day = ts.strftime("%Y-%m-%d") if isinstance(ts, datetime) else str(ts)[:10]
    raw = (doc.get("raw_text") or "")[:80].lower()
    try:
        amount = f"{float(doc.get('amount') or 0):.2f}"
    except (TypeError, ValueError):
        amount = str(doc.get("amount"))
    return f"{doc.get('type')}|{amount}|{day}|{raw}"


def _soft_fingerprint(doc: dict[str, Any]) -> str:
    ts = doc.get("received_at")
    day = ts.strftime("%Y-%m-%d") if isinstance(ts, datetime) else str(ts)[:10]
    label = clean_merchant_label(doc.get("merchant"), fallback=doc.get("raw_text")).lower()
    label = re.sub(r"[^a-z0-9]", "", label)[:18]
    try:
        amount = f"{float(doc.get('amount') or 0):.2f}"
    except (TypeError, ValueError):
        amount = str(doc.get("amount"))
    return f"{doc.get('type')}|{amount}|{day}|{label}"


def review_agent(state: ImportGraphState) -> ImportGraphState:
    """Prepare clarifying questions for low-confidence / flagged rows (Learn About Me)."""
    review = list(state.get("review_queue") or [])
    questions: list[dict[str, Any]] = []
    for item in review[:12]:
        reason = str(item.get("review_reason") or "")
        if reason.startswith("exact_duplicate") or reason.startswith("likely_"):
            continue  # dups are informational, not asked
        merchant = clean_merchant_label(item.get("merchant"), fallback=item.get("raw_text"))
        if not merchant or merchant.lower() in {"unknown", "bank transfer"}:
            continue
        opts = [c for c in CATEGORIES if c != "Other"][:6]
        questions.append(
            {
                "id": f"import_{abs(hash(merchant.lower())) % 10_000_000}",
                "fact_type": "category_intent",
                "key": merchant,
                "priority": 10,
                "prompt": f'Import review: what category is "{merchant}"?',
                "choices": [{"id": c, "label": c} for c in opts],
                "allow_free_text": False,
                "meta": {
                    "amount": item.get("amount"),
                    "raw_text": (item.get("raw_text") or "")[:160],
                    "suggested": item.get("category"),
                    "confidence": item.get("category_confidence"),
                    "review_reason": reason,
                    "source": "import_pipeline",
                },
            }
        )

    stage = "ready_for_review" if review or state.get("needs_mapping_ui") else "complete"
    return {
        **state,
        "stage": stage,
        "review_questions": questions,
        "steps": _append_step(
            state,
            _step(
                "review_agent",
                "ok" if not questions else "warn",
                f"{len(questions)} clarifying questions · {len(review)} rows held for review",
                questions=len(questions),
                held=len(review),
            ),
        ),
    }


def build_import_graph():
    g = StateGraph(ImportGraphState)
    g.add_node("extract", extraction_agent)
    g.add_node("schema_map", schema_mapping_agent)
    g.add_node("categorize", categorization_agent)
    g.add_node("validate", validation_dedup_agent)
    g.add_node("review", review_agent)
    g.set_entry_point("extract")
    g.add_edge("extract", "schema_map")
    g.add_edge("schema_map", "categorize")
    g.add_edge("categorize", "validate")
    g.add_edge("validate", "review")
    g.add_edge("review", END)
    return g.compile()


_compiled = None


def get_import_graph():
    global _compiled
    if _compiled is None:
        _compiled = build_import_graph()
    return _compiled


def run_import_pipeline(
    raw: bytes,
    filename: str,
    *,
    flow: Flow = "statement",
    bank: str | None = None,
    merchant_memory: dict[str, str] | None = None,
    exact_fps: set[str] | None = None,
    soft_fps: set[str] | None = None,
    column_mapping: dict[str, Any] | None = None,
    mongo_db: Any | None = None,
) -> dict[str, Any]:
    """Run the LangGraph import pipeline and return a serializable result."""
    store = get_vector_store(mongo_db)
    graph = get_import_graph()
    initial: ImportGraphState = {
        "raw": raw,
        "filename": filename,
        "flow": flow,
        "bank": bank,
        "stage": "extracting",
        "steps": [],
        "warnings": [],
        "column_mapping": column_mapping or {},
        "needs_mapping_ui": False,
        "mapping_confidence": 0.0,
        "_merchant_memory": merchant_memory or {},
        "_exact_fps": list(exact_fps or []),
        "_soft_fps": list(soft_fps or []),
    }
    # TypedDict doesn't love private keys — pass via object mutation in nodes using state get
    final = graph.invoke(initial)

    # If caller provided confirmed mapping, force it
    if column_mapping:
        final["column_mapping"] = column_mapping
        final["needs_mapping_ui"] = False
        final["mapping_confidence"] = max(float(final.get("mapping_confidence") or 0), 0.95)

    return {
        "protocol": "langgraph-import/v1",
        "flow": flow,
        "stage": final.get("stage"),
        "stages": STAGE_ORDER,
        "file_kind": final.get("file_kind"),
        "sheets": final.get("sheets") or [],
        "headers": final.get("headers") or [],
        "header_signature": final.get("header_signature"),
        "column_mapping": final.get("column_mapping") or {},
        "mapping_confidence": final.get("mapping_confidence") or 0,
        "needs_mapping_ui": bool(final.get("needs_mapping_ui")),
        "profile": final.get("profile") or {},
        "commit_rows": final.get("commit_rows") or [],
        "review_queue": final.get("review_queue") or [],
        "review_questions": final.get("review_questions") or [],
        "anomalies": final.get("anomalies") or [],
        "raw_row_count": len(final.get("raw_rows") or []),
        "steps": final.get("steps") or [],
        "warnings": final.get("warnings") or [],
        "vector_backend": store.backend,
    }


def feedback_document_format(
    store: VectorStore,
    *,
    flow: Flow,
    header_signature: str,
    column_mapping: dict[str, Any],
    bank: str | None = None,
    sheet_purpose: str | None = None,
    filename_hint: str | None = None,
) -> dict[str, Any]:
    return store.upsert_document_format(
        flow=flow,
        header_signature=header_signature,
        column_mapping=column_mapping,
        bank=bank,
        sheet_purpose=sheet_purpose,
        filename_hint=filename_hint,
    )


def feedback_merchant_correction(
    store: VectorStore,
    *,
    description: str,
    category: str,
    merchant_key: str | None = None,
    classification: str | None = None,
) -> dict[str, Any]:
    return store.upsert_merchant_knowledge(
        description=description,
        category=category,
        classification=classification,
        merchant_key=merchant_key,
        confidence=1.0,
        source="user_correction",
    )
