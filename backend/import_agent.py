"""Import agent protocol — staged tools with RAG for PDF/Excel/CSV statements.

Tools (in order):
  1. inspect_file
  2. extract_content
  3. detect_bank
  4. retrieve_templates (RAG)
  5. parse_transactions
  6. rag_categorize
  7. summarize
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from typing import Any, Optional

from import_rag import ImportRAG
from statement_parser import (
    detect_statement_profile,
    enrich_imported_txn,
    parse_statement_csv,
    parse_statement_excel,
    parse_statement_file,
    parse_statement_text,
    _pairs_to_docs,
)


@dataclass
class AgentStep:
    tool: str
    status: str  # ok | warn | error | skip
    detail: str
    evidence: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class ImportAgentResult:
    docs: list[dict[str, Any]]
    profile: dict[str, Any]
    steps: list[AgentStep]
    rag_hits: int = 0
    rag_categorized: int = 0
    warnings: list[str] = field(default_factory=list)

    def to_meta(self) -> dict[str, Any]:
        return {
            "protocol": "import-agent/v1",
            "steps": [s.to_dict() for s in self.steps],
            "rag_hits": self.rag_hits,
            "rag_categorized": self.rag_categorized,
            "warnings": self.warnings,
        }


def _inspect_file(raw: bytes, filename: str) -> AgentStep:
    name = (filename or "upload").lower()
    kind = "unknown"
    if name.endswith(".pdf"):
        kind = "pdf"
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        kind = "excel"
    elif name.endswith((".csv", ".tsv")):
        kind = "csv"
    elif name.endswith(".txt"):
        kind = "text"

    evidence: dict[str, Any] = {
        "filename": filename,
        "bytes": len(raw),
        "kind": kind,
    }

    if kind == "excel":
        try:
            if name.endswith(".xls") and not name.endswith(".xlsx"):
                import xlrd  # type: ignore

                book = xlrd.open_workbook(file_contents=raw)
                evidence["sheets"] = book.nsheets
                evidence["sheet_names"] = list(book.sheet_names())[:8]
            else:
                from openpyxl import load_workbook

                wb = load_workbook(__import__("io").BytesIO(raw), read_only=True, data_only=True)
                try:
                    evidence["sheets"] = len(wb.sheetnames)
                    evidence["sheet_names"] = list(wb.sheetnames)[:8]
                finally:
                    wb.close()
        except Exception as exc:  # noqa: BLE001
            return AgentStep("inspect_file", "warn", f"Excel inspect partial: {exc}", evidence)

    if kind == "pdf":
        try:
            import pdfplumber
            import io

            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                evidence["pages"] = len(pdf.pages)
        except Exception as exc:  # noqa: BLE001
            return AgentStep("inspect_file", "warn", f"PDF inspect partial: {exc}", evidence)

    return AgentStep(
        "inspect_file",
        "ok",
        f"Detected {kind} · {len(raw):,} bytes",
        evidence,
    )


def _extract_pdf_improved(raw: bytes, bank: Optional[str]) -> tuple[list[dict[str, Any]], dict[str, Any], str]:
    """Table-first PDF parse with text fallback; returns docs, stats, peek_text."""
    import io
    import pdfplumber

    table_pairs: list = []
    text_chunks: list[str] = []
    table_count = 0

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                if not table or len(table) < 2:
                    continue
                table_count += 1
                cleaned = [
                    [("" if c is None else str(c).replace("\n", " ").strip()) for c in row]
                    for row in table
                    if row and any(c is not None and str(c).strip() for c in row)
                ]
                if len(cleaned) < 2:
                    continue
                from statement_parser import _find_header_row, _rows_to_csv_text

                header_idx = _find_header_row(cleaned)
                pairs = parse_statement_csv(
                    _rows_to_csv_text(cleaned[header_idx:]), bank=bank
                )
                table_pairs.extend(pairs)

            page_text = page.extract_text() or ""
            if page_text.strip():
                text_chunks.append(page_text)

    peek = "\n".join(text_chunks)[:12000]
    text_pairs = parse_statement_text(peek, bank=bank) if peek.strip() else []

    # Prefer the richer extraction path
    if len(table_pairs) >= max(3, len(text_pairs)):
        chosen = "tables"
        pairs = table_pairs
    elif text_pairs:
        chosen = "text"
        pairs = text_pairs
    else:
        chosen = "tables" if table_pairs else "none"
        pairs = table_pairs or text_pairs

    # Deduplicate identical fingerprints within this extract
    seen: set[str] = set()
    unique = []
    for txn, received_at in pairs:
        key = f"{txn.get('type')}|{txn.get('amount')}|{received_at.date().isoformat()}|{str(txn.get('raw_text') or '')[:40]}"
        if key in seen:
            continue
        seen.add(key)
        unique.append((txn, received_at))

    stats = {
        "tables_found": table_count,
        "table_rows": len(table_pairs),
        "text_rows": len(text_pairs),
        "chosen": chosen,
        "unique_rows": len(unique),
        "pages_with_text": len(text_chunks),
    }
    return unique, stats, peek


def _extract_excel_improved(
    raw: bytes, filename: str, bank: Optional[str]
) -> tuple[list, dict[str, Any]]:
    """Try all sheets; keep the sheet with the most parsed transactions."""
    name = (filename or "").lower()
    sheet_scores: list[dict[str, Any]] = []
    best_pairs: list = []

    try:
        if name.endswith(".xls") and not name.endswith(".xlsx"):
            import xlrd  # type: ignore

            book = xlrd.open_workbook(file_contents=raw)
            for i in range(book.nsheets):
                sheet = book.sheet_by_index(i)
                rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
                from statement_parser import _find_header_row, _rows_to_csv_text

                if not rows:
                    continue
                header_idx = _find_header_row(rows)
                table = [r for r in rows[header_idx:] if any(c is not None and str(c).strip() for c in r)]
                pairs = parse_statement_csv(_rows_to_csv_text(table), bank=bank) if len(table) >= 2 else []
                sheet_scores.append(
                    {"name": sheet.name, "rows": len(rows), "parsed": len(pairs)}
                )
                if len(pairs) > len(best_pairs):
                    best_pairs = pairs
        else:
            import io
            from openpyxl import load_workbook
            from statement_parser import _find_header_row, _rows_to_csv_text

            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = [list(row) for row in ws.iter_rows(values_only=True) if row]
                    if not rows:
                        continue
                    header_idx = _find_header_row(rows)
                    table = [
                        r for r in rows[header_idx:] if any(c is not None and str(c).strip() for c in r)
                    ]
                    pairs = (
                        parse_statement_csv(_rows_to_csv_text(table), bank=bank)
                        if len(table) >= 2
                        else []
                    )
                    sheet_scores.append(
                        {"name": sheet_name, "rows": len(rows), "parsed": len(pairs)}
                    )
                    if len(pairs) > len(best_pairs):
                        best_pairs = pairs
            finally:
                wb.close()
    except Exception:
        # Fall back to original single-sheet parser
        best_pairs = parse_statement_excel(raw, bank=bank, filename=name)
        sheet_scores.append({"name": "(fallback)", "parsed": len(best_pairs)})

    stats = {
        "sheets": sheet_scores,
        "best_parsed": len(best_pairs),
        "best_sheet": max(sheet_scores, key=lambda s: s.get("parsed", 0))["name"]
        if sheet_scores
        else None,
    }
    return best_pairs, stats


def run_import_agent(
    raw: bytes,
    filename: str,
    *,
    bank: Optional[str] = None,
    fmt: str = "auto",
    rag: ImportRAG | None = None,
) -> ImportAgentResult:
    steps: list[AgentStep] = []
    warnings: list[str] = []
    rag = rag or ImportRAG()
    if len(rag) == 0:
        rag.seed_templates()

    # 1) inspect
    inspect = _inspect_file(raw, filename)
    steps.append(inspect)
    kind = str(inspect.evidence.get("kind") or "unknown")

    # 2) extract + 3/4 detect with RAG
    peek = ""
    try:
        peek = raw[:8000].decode("utf-8-sig", errors="ignore")
    except Exception:  # noqa: BLE001
        peek = ""

    pairs: list = []
    extract_stats: dict[str, Any] = {}

    if kind == "pdf" or (filename or "").lower().endswith(".pdf") or fmt == "pdf":
        # Tentative bank from filename / weak peek
        early = detect_statement_profile(f"{filename}\n{peek}", filename)
        bank_guess = bank or early.get("bank")
        pairs, extract_stats, peek = _extract_pdf_improved(raw, bank_guess)
        steps.append(
            AgentStep(
                "extract_content",
                "ok" if pairs else "warn",
                f"PDF extract via {extract_stats.get('chosen')} · {extract_stats.get('unique_rows', 0)} rows",
                extract_stats,
            )
        )
        if not pairs:
            warnings.append("PDF yielded no rows — scanned PDFs need OCR (not enabled)")
    elif kind == "excel" or (filename or "").lower().endswith((".xlsx", ".xls")):
        early = detect_statement_profile(f"{filename}\n{peek}", filename)
        bank_guess = bank or early.get("bank")
        pairs, extract_stats = _extract_excel_improved(raw, filename, bank_guess)
        steps.append(
            AgentStep(
                "extract_content",
                "ok" if pairs else "warn",
                f"Excel sheet '{extract_stats.get('best_sheet')}' · {extract_stats.get('best_parsed', 0)} rows",
                extract_stats,
            )
        )
        if not pairs:
            warnings.append("Excel sheets parsed to zero transactions")
    else:
        try:
            docs, profile = parse_statement_file(raw, filename, bank=bank, fmt=fmt)
            steps.append(
                AgentStep(
                    "extract_content",
                    "ok" if docs else "warn",
                    f"Text/CSV parse · {len(docs)} rows",
                    {"rows": len(docs)},
                )
            )
            # Continue with RAG categorize on docs path
            return _finish_with_rag(
                docs,
                profile,
                steps,
                warnings,
                rag,
                bank=bank,
            )
        except Exception as exc:  # noqa: BLE001
            steps.append(AgentStep("extract_content", "error", str(exc), {}))
            raise

    # Detect bank using extracted peek + RAG templates
    sample_for_detect = f"{filename}\n{peek[:8000]}"
    profile = detect_statement_profile(sample_for_detect, filename)
    rag_bank = rag.suggest_bank(sample_for_detect)
    rag_hits = 0
    if rag_bank:
        rag_hits += 1
        if not bank and (
            profile.get("confidence") in {"low", "medium", None}
            or not profile.get("bank")
        ):
            # Boost with RAG when heuristic is weak
            if rag_bank["confidence"] >= 0.15:
                profile = {**profile, "bank": rag_bank["bank"], "rag_bank": True}
        steps.append(
            AgentStep(
                "retrieve_templates",
                "ok",
                f"RAG bank hint: {rag_bank['bank']} ({rag_bank['confidence']:.2f})",
                rag_bank,
            )
        )
    else:
        steps.append(
            AgentStep("retrieve_templates", "skip", "No strong template match", {})
        )

    bank_name = bank or profile.get("bank")
    default_card = profile.get("default_card_type") or "bank_account"
    if bank_name == "ICICI Bank" and profile.get("is_credit_card_statement"):
        default_card = "credit_card"

    steps.append(
        AgentStep(
            "detect_bank",
            "ok",
            f"{bank_name} · card={default_card} · confidence={profile.get('confidence')}",
            {
                "bank": bank_name,
                "is_credit_card_statement": profile.get("is_credit_card_statement"),
                "confidence": profile.get("confidence"),
            },
        )
    )

    # Re-parse with confirmed bank if we used a weak early guess
    if bank_name and kind == "pdf" and pairs:
        # Enrich existing pairs with bank name later via _pairs_to_docs
        pass
    elif bank_name and kind == "excel" and not pairs:
        pairs, extract_stats = _extract_excel_improved(raw, filename, bank_name)

    docs = _pairs_to_docs(pairs, default_card_type=default_card, bank=bank_name)
    steps.append(
        AgentStep(
            "parse_transactions",
            "ok" if docs else "warn",
            f"Normalized {len(docs)} transaction docs",
            {"count": len(docs)},
        )
    )

    profile = {
        **profile,
        "bank": bank_name,
        "default_card_type": default_card,
    }
    return _finish_with_rag(docs, profile, steps, warnings, rag, bank=bank_name, rag_hits=rag_hits)


def _finish_with_rag(
    docs: list[dict[str, Any]],
    profile: dict[str, Any],
    steps: list[AgentStep],
    warnings: list[str],
    rag: ImportRAG,
    *,
    bank: Optional[str] = None,
    rag_hits: int = 0,
) -> ImportAgentResult:
    bank_name = bank or profile.get("bank")
    categorized = 0
    hit_total = rag_hits

    for doc in docs:
        if doc.get("category") and doc.get("category") != "Other":
            continue
        narration = f"{doc.get('merchant') or ''} {doc.get('raw_text') or ''}"
        suggestion = rag.suggest_category(narration, bank=bank_name)
        if not suggestion:
            continue
        hit_total += 1
        doc["category"] = suggestion["category"]
        doc["category_source"] = f"rag:{suggestion['source']}"
        doc["category_confidence"] = suggestion["confidence"]
        categorized += 1

    steps.append(
        AgentStep(
            "rag_categorize",
            "ok" if categorized else "skip",
            f"RAG filled {categorized} categories from similar past merchants",
            {"categorized": categorized, "index_size": len(rag)},
        )
    )

    # Ensure enrich still applied (idempotent for missing fields)
    default_card = profile.get("default_card_type") or "bank_account"
    for doc in docs:
        enrich_imported_txn(doc, default_card_type=default_card)

    steps.append(
        AgentStep(
            "summarize",
            "ok",
            f"Ready to import {len(docs)} rows · RAG hits {hit_total}",
            {"docs": len(docs), "rag_hits": hit_total, "rag_categorized": categorized},
        )
    )

    return ImportAgentResult(
        docs=docs,
        profile=profile,
        steps=steps,
        rag_hits=hit_total,
        rag_categorized=categorized,
        warnings=warnings,
    )
