"""INDmoney CSV/Excel import — parse, map columns, validate, upsert."""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import load_workbook

# Target schema fields the user can map onto
TARGET_FIELDS = [
    {"key": "name", "label": "Instrument name", "required": True},
    {"key": "asset_class", "label": "Asset class", "required": False},
    {"key": "invested", "label": "Invested amount", "required": False},
    {"key": "current_value", "label": "Current value", "required": False},
    {"key": "units", "label": "Units / quantity", "required": False},
    {"key": "as_of_date", "label": "As-of / trade date", "required": False},
]

# Header synonyms → target field (lowercase, stripped)
HEADER_ALIASES: dict[str, list[str]] = {
    "name": [
        "name",
        "instrument",
        "instrument name",
        "scheme",
        "scheme name",
        "stock",
        "stock name",
        "security",
        "security name",
        "fund",
        "fund name",
        "holding",
        "scrip",
        "symbol",
        "ticker",
        "product",
        "asset name",
    ],
    "asset_class": [
        "asset class",
        "assetclass",
        "category",
        "asset type",
        "type",
        "segment",
        "product type",
        "class",
    ],
    "invested": [
        "invested",
        "invested amount",
        "investment",
        "investment amount",
        "cost",
        "cost value",
        "cost amount",
        "purchase value",
        "buy value",
        "amount invested",
        "invested value",
        "total cost",
        "avg cost",
    ],
    "current_value": [
        "current value",
        "currentvalue",
        "market value",
        "marketvalue",
        "present value",
        "value",
        "nav value",
        "current amount",
        "valuation",
        "total value",
        "mkt value",
    ],
    "units": [
        "units",
        "quantity",
        "qty",
        "shares",
        "holding units",
        "no of units",
        "number of units",
    ],
    "as_of_date": [
        "date",
        "as of",
        "as of date",
        "as_of",
        "asof",
        "trade date",
        "transaction date",
        "nav date",
        "updated on",
        "valuation date",
    ],
}


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def suggest_mapping(headers: list[str]) -> dict[str, str | None]:
    """Map each target field → source header (or None)."""
    normalized = {_norm_header(h): h for h in headers}
    mapping: dict[str, str | None] = {f["key"]: None for f in TARGET_FIELDS}
    used: set[str] = set()

    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized and normalized[alias] not in used:
                mapping[field] = normalized[alias]
                used.add(normalized[alias])
                break
            # fuzzy contains
            for nh, original in normalized.items():
                if original in used:
                    continue
                if alias in nh or nh in alias:
                    mapping[field] = original
                    used.add(original)
                    break
            if mapping[field]:
                break
    return mapping


def _parse_amount(raw: Any) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text or text in {"-", "—", "NA", "N/A", "null"}:
        return None
    text = text.replace(",", "").replace("₹", "").replace("Rs.", "").replace("INR", "")
    text = text.replace("(", "-").replace(")", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def _parse_date(raw: Any) -> Optional[str]:
    """Return ISO date YYYY-MM-DD or None."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    text = str(raw).strip()
    if not text:
        return None
    # Excel serial sometimes arrives as float string
    try:
        as_float = float(text)
        if 20000 < as_float < 80000:
            # Excel epoch roughly
            from datetime import date, timedelta

            base = date(1899, 12, 30)
            return (base + timedelta(days=int(as_float))).isoformat()
    except ValueError:
        pass

    for fmt in (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d-%b-%Y",
        "%d %b %Y",
        "%d-%b-%y",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def read_tabular_file(raw: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (headers, list of row dicts keyed by header)."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xls")):
        return _read_excel(raw)
    return _read_csv(raw)


def _read_csv(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = raw.decode("utf-8-sig", errors="replace")
    # Detect delimiter
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [h for h in (reader.fieldnames or []) if h]
    rows: list[dict[str, Any]] = []
    for row in reader:
        cleaned = {k: (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
        if any(str(v or "").strip() for v in cleaned.values()):
            rows.append(cleaned)
    return headers, rows


def _read_excel(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(header_row)]
    # Drop fully empty header names at end
    while headers and (not headers[-1] or headers[-1].startswith("col_")):
        # keep named cols; stop trimming once we hit content names mid-row
        if headers[-1].startswith("col_"):
            headers.pop()
        else:
            break
    out: list[dict[str, Any]] = []
    for cells in rows_iter:
        row = {}
        empty = True
        for i, h in enumerate(headers):
            val = cells[i] if i < len(cells) else None
            if val is not None and str(val).strip() != "":
                empty = False
            row[h] = val
        if not empty:
            out.append(row)
    return headers, out


def apply_mapping(
    rows: list[dict[str, Any]],
    mapping: dict[str, str | None],
) -> dict[str, Any]:
    """Validate mapped rows and return preview payload."""
    valid: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    duplicate_keys: set[str] = set()

    name_col = mapping.get("name")
    if not name_col:
        return {
            "ok": False,
            "error": "Instrument name column is required",
            "valid": [],
            "errors": [],
            "summary": {"valid": 0, "errors": 0, "duplicates_in_file": 0},
        }

    for idx, row in enumerate(rows, start=2):  # 1-indexed + header
        name_raw = row.get(name_col)
        name = str(name_raw or "").strip()
        if not name:
            errors.append({"row": idx, "issue": "Missing instrument name", "raw": row})
            continue

        invested = _parse_amount(row.get(mapping["invested"])) if mapping.get("invested") else None
        current = (
            _parse_amount(row.get(mapping["current_value"]))
            if mapping.get("current_value")
            else None
        )
        units = _parse_amount(row.get(mapping["units"])) if mapping.get("units") else None

        as_of: Optional[str] = None
        if mapping.get("as_of_date"):
            raw_date = row.get(mapping["as_of_date"])
            as_of = _parse_date(raw_date)
            if raw_date not in (None, "") and as_of is None:
                errors.append(
                    {
                        "row": idx,
                        "issue": f"Malformed date: {raw_date!r}",
                        "instrument": name,
                    }
                )
                continue

        asset_class = "Other Investments"
        if mapping.get("asset_class"):
            ac = row.get(mapping["asset_class"])
            if ac is not None and str(ac).strip():
                asset_class = str(ac).strip()[:60]

        if invested is None and current is None:
            errors.append(
                {
                    "row": idx,
                    "issue": "Need invested amount and/or current value",
                    "instrument": name,
                }
            )
            continue

        # Defaults: if one amount missing, mirror the other
        if invested is None:
            invested = current or 0.0
        if current is None:
            current = invested

        if invested < 0 or current < 0:
            errors.append(
                {
                    "row": idx,
                    "issue": "Negative amounts not allowed",
                    "instrument": name,
                }
            )
            continue

        instrument_key = _instrument_key(name, as_of)
        if instrument_key in seen_keys:
            duplicate_keys.add(instrument_key)
        seen_keys.add(instrument_key)

        valid.append(
            {
                "name": name[:80],
                "asset_class": asset_class,
                "invested": round(float(invested), 2),
                "current_value": round(float(current), 2),
                "units": round(float(units), 4) if units is not None else None,
                "as_of_date": as_of,
                "instrument_key": instrument_key,
            }
        )

    return {
        "ok": True,
        "valid": valid,
        "errors": errors,
        "summary": {
            "valid": len(valid),
            "errors": len(errors),
            "duplicates_in_file": len(duplicate_keys),
            "total_invested": round(sum(v["invested"] for v in valid), 2),
            "total_current": round(sum(v["current_value"] for v in valid), 2),
        },
    }


def _instrument_key(name: str, as_of: str | None) -> str:
    base = re.sub(r"\s+", " ", name.strip().lower())
    return f"{base}|{as_of or 'latest'}"


def upsert_holdings(
    portfolio_collection: Any,
    holdings: list[dict[str, Any]],
    *,
    source: str = "csv_import",
    user_id: Any = None,
) -> dict[str, Any]:
    """Upsert by instrument_key (name + as_of_date). Tags source + import_date."""
    now = datetime.now(timezone.utc)
    inserted = 0
    updated = 0

    for h in holdings:
        key = h.get("instrument_key") or _instrument_key(h["name"], h.get("as_of_date"))
        doc: dict[str, Any] = {
            "name": h["name"],
            "asset_class": h.get("asset_class") or "Other Investments",
            "invested": float(h["invested"]),
            "current_value": float(h["current_value"]),
            "units": h.get("units"),
            "as_of_date": h.get("as_of_date"),
            "instrument_key": key,
            "source": source,
            "import_date": now,
            "updated_at": now,
            # Reserved for Account Aggregator later
            "external_id": h.get("external_id"),
            "provider": h.get("provider") or "indmoney",
        }
        if user_id is not None:
            doc["user_id"] = user_id
        lookup: dict[str, Any] = {"instrument_key": key}
        if user_id is not None:
            lookup["user_id"] = user_id
        existing = portfolio_collection.find_one(lookup)
        if existing:
            portfolio_collection.update_one({"_id": existing["_id"]}, {"$set": doc})
            updated += 1
        else:
            # Also match legacy docs that only have name (no instrument_key)
            legacy_q: dict[str, Any] = {"name": h["name"], "instrument_key": {"$exists": False}}
            if user_id is not None:
                legacy_q["user_id"] = user_id
            legacy = portfolio_collection.find_one(legacy_q)
            if legacy and not h.get("as_of_date"):
                portfolio_collection.update_one({"_id": legacy["_id"]}, {"$set": doc})
                updated += 1
            else:
                portfolio_collection.insert_one(doc)
                inserted += 1

    return {"inserted": inserted, "updated": updated, "total": inserted + updated}
